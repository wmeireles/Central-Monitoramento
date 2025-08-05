import csv
import logging
from datetime import datetime, timedelta
from typing import List, Dict
# import matplotlib.pyplot as plt
# import matplotlib.dates as mdates
from database import MonitoringDB

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class ReportGenerator:
    def __init__(self, db: MonitoringDB):
        self.db = db
    
    def generate_csv_report(self, site_name: str = None, days: int = 30) -> str:
        """Gera relatório em CSV"""
        filename = f"report_{site_name or 'all_sites'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['timestamp', 'site_name', 'url', 'status_code', 
                         'response_time', 'ssl_days_remaining', 'is_up', 'error_message']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            
            # Buscar dados do banco
            if site_name:
                data = self.db.get_site_history(site_name, limit=days * 24 * 12)  # Assumindo check a cada 5min
            else:
                # Implementar método para buscar todos os sites
                data = []
            
            for row in data:
                writer.writerow(row)
        
        logging.info(f"Relatório CSV gerado: {filename}")
        return filename
    
    def generate_excel_report(self, site_name: str = None, days: int = 30) -> str:
        """Gera relatório em Excel com formatação"""
        if not EXCEL_AVAILABLE:
            logging.error("openpyxl não está disponível para gerar relatórios Excel")
            return None
        
        filename = f"report_{site_name or 'all_sites'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Monitoring Report"
        
        # Cabeçalhos
        headers = ['Timestamp', 'Site', 'URL', 'Status Code', 'Response Time (ms)', 
                  'SSL Days', 'Status', 'Error Message']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Dados
        if site_name:
            data = self.db.get_site_history(site_name, limit=days * 24 * 12)
        else:
            data = []
        
        for row_idx, row_data in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=row_data['timestamp'])
            ws.cell(row=row_idx, column=2, value=row_data['site_name'])
            ws.cell(row=row_idx, column=3, value=row_data['url'])
            ws.cell(row=row_idx, column=4, value=row_data['status_code'])
            ws.cell(row=row_idx, column=5, value=row_data['response_time'])
            ws.cell(row=row_idx, column=6, value=row_data['ssl_days_remaining'])
            ws.cell(row=row_idx, column=7, value="UP" if row_data['is_up'] else "DOWN")
            ws.cell(row=row_idx, column=8, value=row_data['error_message'])
            
            # Colorir linha baseado no status
            if not row_data['is_up']:
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                    )
        
        wb.save(filename)
        logging.info(f"Relatório Excel gerado: {filename}")
        return filename
    
    def generate_uptime_chart(self, site_name: str, days: int = 7) -> str:
        """Gera relatório de uptime em texto (matplotlib não disponível)"""
        data = self.db.get_site_history(site_name, limit=days * 24 * 12)
        
        if not data:
            logging.warning(f"Nenhum dado encontrado para {site_name}")
            return None
        
        # Gerar relatório em texto
        filename = f"chart_{site_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(f"Relatório de Uptime - {site_name}\n")
            f.write(f"Período: Últimos {days} dias\n")
            f.write("=" * 50 + "\n\n")
            
            # Estatísticas
            stats = self.calculate_uptime_percentage(site_name, days)
            f.write(f"Uptime: {stats['uptime_percentage']:.2f}%\n")
            f.write(f"Tempo médio de resposta: {stats['avg_response_time']:.2f}ms\n")
            f.write(f"Total de verificações: {stats['total_checks']}\n")
            f.write(f"Verificações com sucesso: {stats['successful_checks']}\n")
            f.write(f"Verificações com falha: {stats['failed_checks']}\n\n")
            
            # Histórico recente
            f.write("Histórico (últimas 20 verificações):\n")
            f.write("-" * 50 + "\n")
            
            for row in data[:20]:
                status = "UP" if row['is_up'] else "DOWN"
                timestamp = row['timestamp']
                response_time = row['response_time'] or 0
                f.write(f"{timestamp} | {status:4} | {response_time:6.0f}ms\n")
        
        logging.info(f"Relatório de texto gerado: {filename}")
        return filename
    
    def calculate_uptime_percentage(self, site_name: str, days: int = 30) -> Dict:
        """Calcula estatísticas de uptime"""
        data = self.db.get_site_history(site_name, limit=days * 24 * 12)
        
        if not data:
            return {'uptime_percentage': 0, 'total_checks': 0, 'successful_checks': 0}
        
        total_checks = len(data)
        successful_checks = sum(1 for row in data if row['is_up'])
        uptime_percentage = (successful_checks / total_checks) * 100
        
        # Calcular tempo médio de resposta
        response_times = [row['response_time'] for row in data if row['response_time']]
        avg_response_time = sum(response_times) / len(response_times) if response_times else 0
        
        return {
            'uptime_percentage': round(uptime_percentage, 2),
            'total_checks': total_checks,
            'successful_checks': successful_checks,
            'failed_checks': total_checks - successful_checks,
            'avg_response_time': round(avg_response_time, 2)
        }